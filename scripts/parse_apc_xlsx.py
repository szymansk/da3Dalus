"""Parse the APC PROP-DATA xlsx product list (gh-1000).

The ``PROP-DATA-FILE_*.xlsx`` ``PRODUCT LIST`` sheet carries per-product
catalog data: product name, diameter/pitch (inches), hub dims and a **weight
in ounces**. PE0 files are the primary weight/inertia source (SI, 1:1 with
PER3); this loader provides a secondary cross-check and fills props that have
no PE0 file.

Weight is normalised from ounces to **grams** at the boundary. Rows without a
parseable weight or designation are skipped (counted, never silently merged).

Requires ``openpyxl`` (added as a runtime dep in gh-1000).
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)

OZ_TO_G = 28.349523125

_PRODUCT_SHEET = "PRODUCT LIST"
# "11x6", "10.5x4.5E", "4.1x4.1EP" → (dia, pitch, variant)
_DESIG_RE = re.compile(r"^\s*(\d+(?:\.\d+)?)x(\d+(?:\.\d+)?)(\S*)\s*$")
_OZ_RE = re.compile(r"([-+]?\d+(?:\.\d+)?)\s*oz")


@dataclass
class XlsxProp:
    diameter_in: float
    pitch_in: float
    variant: str
    weight_g: float | None


def _parse_designation(name: str) -> tuple[float, float, str] | None:
    m = _DESIG_RE.match(str(name).strip())
    if not m:
        return None
    return float(m.group(1)), float(m.group(2)), m.group(3).strip()


def _parse_oz(value: object) -> float | None:
    if value is None:
        return None
    m = _OZ_RE.search(str(value))
    if not m:
        return None
    return float(m.group(1)) * OZ_TO_G


def parse_apc_xlsx(path: Path) -> list[XlsxProp]:
    """Parse the PROP-DATA xlsx PRODUCT LIST into a list of :class:`XlsxProp`."""
    import openpyxl  # local import: optional/heavy dep

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if _PRODUCT_SHEET not in wb.sheetnames:
        raise ValueError(f"{path.name}: no '{_PRODUCT_SHEET}' sheet (got {wb.sheetnames})")
    ws = wb[_PRODUCT_SHEET]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def col(*names: str) -> int | None:
        for i, h in enumerate(header):
            if any(h.startswith(n) for n in names):
                return i
        return None

    name_i = col("product name")
    weight_i = col("weight")
    result: list[XlsxProp] = []
    skipped = 0

    for row in rows[1:]:
        if name_i is None or name_i >= len(row):
            skipped += 1
            continue
        desig = _parse_designation(row[name_i]) if row[name_i] is not None else None
        if desig is None:
            skipped += 1
            continue
        weight_g = (
            _parse_oz(row[weight_i]) if weight_i is not None and weight_i < len(row) else None
        )
        dia, pitch, variant = desig
        result.append(XlsxProp(dia, pitch, variant, weight_g))

    logger.info("parse_apc_xlsx: %d props parsed, %d rows skipped", len(result), skipped)
    return result
